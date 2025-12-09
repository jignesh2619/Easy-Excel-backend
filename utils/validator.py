"""
Data Validation Utilities

Validates uploaded files and data before processing:
- File format (CSV/XLSX only)
- Sheet not empty
- Column availability before applying operations
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook


class DataValidator:
    """Validates files and data operations"""
    
    ALLOWED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    
    @staticmethod
    def validate_file_format(filename: str) -> Tuple[bool, Optional[str]]:
        """
        Validate file has allowed extension
        
        Args:
            filename: Name of the file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        file_ext = Path(filename).suffix.lower()
        if file_ext not in DataValidator.ALLOWED_EXTENSIONS:
            return False, f"File format '{file_ext}' not supported. Allowed formats: CSV, XLSX, XLS"
        return True, None
    
    @staticmethod
    def validate_file_size(file_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validate file size is within limits
        
        Args:
            file_path: Path to file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        file_size = Path(file_path).stat().st_size
        if file_size > DataValidator.MAX_FILE_SIZE:
            size_mb = file_size / (1024 * 1024)
            max_mb = DataValidator.MAX_FILE_SIZE / (1024 * 1024)
            return False, f"File size {size_mb:.2f}MB exceeds maximum of {max_mb}MB"
        return True, None
    
    @staticmethod
    def validate_file_exists(file_path: str) -> Tuple[bool, Optional[str]]:
        """
        Check if file exists
        
        Args:
            file_path: Path to file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not Path(file_path).exists():
            return False, f"File not found: {file_path}"
        return True, None
    
    @staticmethod
    def validate_sheet_not_empty(file_path: str, sheet_name: Optional[str] = None) -> Tuple[bool, Optional[str], Optional[pd.DataFrame]]:
        """
        Validate that the sheet contains data
        
        Args:
            file_path: Path to Excel/CSV file
            sheet_name: Name of sheet (for Excel files), None for CSV or first sheet
            
        Returns:
            Tuple of (is_valid, error_message, dataframe)
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                # Try different encodings for CSV files
                encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
                df = None
                last_error = None
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding, on_bad_lines='skip')
                        break
                    except UnicodeDecodeError as e:
                        last_error = e
                        continue
                    except Exception as e:
                        last_error = e
                        continue
                
                if df is None:
                    return False, f"Error reading CSV file. Please ensure the file is properly formatted. Details: {str(last_error)}", None
                    
            elif file_ext in ['.xlsx', '.xls']:
                # Always specify sheet_name to avoid getting a dict
                # If sheet_name is None, use 0 to get first sheet
                try:
                    if sheet_name is None:
                        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl' if file_ext == '.xlsx' else None)
                    else:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl' if file_ext == '.xlsx' else None)
                except Exception as e:
                    # Try without engine specification for .xls files
                    if file_ext == '.xls':
                        try:
                            if sheet_name is None:
                                df = pd.read_excel(file_path, sheet_name=0)
                            else:
                                df = pd.read_excel(file_path, sheet_name=sheet_name)
                        except Exception as e2:
                            return False, f"Error reading Excel file. The file may be corrupted or in an unsupported format. Details: {str(e2)}", None
                    else:
                        return False, f"Error reading Excel file. The file may be corrupted or in an unsupported format. Details: {str(e)}", None
                
                # Check if we got a dict (shouldn't happen with sheet_name specified, but double-check)
                if isinstance(df, dict):
                    # If it's a dict, get the first sheet
                    if len(df) > 0:
                        df = list(df.values())[0]
                    else:
                        return False, "File has no sheets", None
            else:
                return False, "Unsupported file format", None
            
            # Ensure we have a DataFrame
            if not isinstance(df, pd.DataFrame):
                return False, f"Unexpected data type: {type(df)}. Expected DataFrame.", None
            
            if df.empty:
                return False, "File is empty - no data found. Please ensure your file contains data rows.", None
            
            if len(df.columns) == 0:
                return False, "File has no columns. Please ensure your file has a header row.", None
            
            return True, None, df
            
        except pd.errors.EmptyDataError:
            return False, "File appears to be empty or corrupted. Please check the file and try again.", None
        except pd.errors.ParserError as e:
            return False, f"Error parsing file structure. Please ensure the file is properly formatted. Details: {str(e)}", None
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            return False, f"Error reading file: {str(e)}. Please ensure the file is not corrupted and is in a supported format (CSV, XLSX, or XLS).", None
    
    @staticmethod
    def validate_columns_exist(df: pd.DataFrame, required_columns: List[str]) -> Tuple[bool, Optional[str], List[str]]:
        """
        Validate that required columns exist in dataframe
        
        Args:
            df: Dataframe to check
            required_columns: List of column names required
            
        Returns:
            Tuple of (is_valid, error_message, missing_columns)
        """
        if not required_columns:
            return True, None, []
        
        existing_columns = set(df.columns.str.strip())
        required_set = set(col.strip() for col in required_columns)
        missing_columns = list(required_set - existing_columns)
        
        if missing_columns:
            available_cols = ", ".join(existing_columns)
            return False, f"Required columns not found: {', '.join(missing_columns)}. Available columns: {available_cols}", missing_columns
        
        return True, None, []
    
    @staticmethod
    def get_available_columns(file_path: str, sheet_name: Optional[str] = None) -> Tuple[bool, Optional[str], List[str]]:
        """
        Get list of available columns in file
        
        Args:
            file_path: Path to Excel/CSV file
            sheet_name: Name of sheet (for Excel files)
            
        Returns:
            Tuple of (success, error_message, column_list)
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(file_path, nrows=1)  # Read just headers
            elif file_ext in ['.xlsx', '.xls']:
                # Always specify sheet_name to avoid getting a dict
                if sheet_name is None:
                    df = pd.read_excel(file_path, sheet_name=0, nrows=1)
                else:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
                
                # Check if we got a dict (shouldn't happen, but double-check)
                if isinstance(df, dict):
                    if len(df) > 0:
                        df = list(df.values())[0]
                    else:
                        return False, "File has no sheets", []
            else:
                return False, "Unsupported file format", []
            
            # Ensure we have a DataFrame
            if not isinstance(df, pd.DataFrame):
                return False, f"Unexpected data type: {type(df)}. Expected DataFrame.", []
            
            return True, None, list(df.columns)
            
        except Exception as e:
            return False, f"Error reading file: {str(e)}", []
    
    @staticmethod
    def get_sheet_names(file_path: str) -> Tuple[bool, Optional[str], List[str]]:
        """
        Get list of sheet names in Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (success, error_message, sheet_names)
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            if file_ext == '.csv':
                return True, None, ['Sheet1']  # CSV has no sheets
            
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return True, None, sheet_names
            
        except Exception as e:
            return False, f"Error reading Excel file: {str(e)}", []
    
    @staticmethod
    def validate_complete_file(file_path: str, filename: str) -> Tuple[bool, Optional[str], Optional[pd.DataFrame]]:
        """
        Complete validation of uploaded file
        
        Args:
            file_path: Path to file
            filename: Original filename
            
        Returns:
            Tuple of (is_valid, error_message, dataframe)
        """
        # Check file exists
        is_valid, error = DataValidator.validate_file_exists(file_path)
        if not is_valid:
            return False, error, None
        
        # Check file format
        is_valid, error = DataValidator.validate_file_format(filename)
        if not is_valid:
            return False, error, None
        
        # Check file size
        is_valid, error = DataValidator.validate_file_size(file_path)
        if not is_valid:
            return False, error, None
        
        # Check sheet not empty
        is_valid, error, df = DataValidator.validate_sheet_not_empty(file_path)
        if not is_valid:
            return False, error, None
        
        return True, None, df


