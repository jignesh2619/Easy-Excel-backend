"""
Openpyxl-based Excel Writer

Uses openpyxl for writing Excel files with formula support.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Any
from pathlib import Path


class OpenpyxlWriter:
    """Excel writer using openpyxl engine (supports formulas)"""
    
    def __init__(self, output_path: str):
        """
        Initialize writer
        
        Args:
            output_path: Path to output Excel file
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
    
    def write(self, df: pd.DataFrame, sheet_name: str = 'Sheet1',
              formatting_rules: Optional[List[Dict]] = None) -> str:
        """
        Write dataframe to Excel with formatting
        
        Args:
            df: DataFrame to write
            sheet_name: Name of the sheet
            formatting_rules: List of formatting rules to apply
        
        Returns:
            Path to written file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting rules
        if formatting_rules:
            self._apply_formatting_rules(ws, df, formatting_rules)
        
        # Auto-adjust column widths
        for col_idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        wb.save(self.output_path)
        return str(self.output_path)
    
    def _apply_formatting_rules(self, worksheet, df: pd.DataFrame, rules: List[Dict]):
        """Apply formatting rules to cells"""
        for rule in rules:
            if rule.get("type") != "format":
                continue
            
            formatting = rule.get("formatting", {})
            range_info = rule.get("range", {})
            
            # Build font
            font = Font()
            if formatting.get("bold"):
                font.bold = True
            if formatting.get("italic"):
                font.italic = True
            if formatting.get("text_color"):
                font.color = formatting['text_color']
            
            # Build fill
            fill = None
            if formatting.get("bg_color"):
                fill = PatternFill(start_color=formatting['bg_color'], 
                                 end_color=formatting['bg_color'], 
                                 fill_type='solid')
            
            # Apply to range
            if "cells" in range_info:
                cells = range_info["cells"]
                for cell_info in cells:
                    row_idx = cell_info.get("row", 0)
                    col_name = cell_info.get("column")
                    if col_name in df.columns and 0 <= row_idx < len(df):
                        col_idx = list(df.columns).index(col_name) + 1
                        excel_row = row_idx + 2  # +1 for header, +1 for 1-indexed
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell.font = font
                        if fill:
                            cell.fill = fill

