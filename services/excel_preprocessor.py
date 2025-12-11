"""
Excel Preprocessor Service

Preprocesses Excel files before analysis.
Handles merged cells, headers, formatting issues.
Simplified version adapted from excelaiagent's dismantle_excel.py
"""

import logging
import openpyxl
import pandas as pd
from typing import Optional, Dict
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelPreprocessor:
    """Preprocesses Excel files for better analysis"""
    
    def preprocess(self, file_path: str, output_path: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        Preprocess Excel: unmerge, fix headers, normalize
        
        Args:
            file_path: Path to input Excel file
            output_path: Optional path to save preprocessed file (if None, returns DataFrame only)
            
        Returns:
            Preprocessed DataFrame or None if preprocessing fails
        """
        try:
            # Load with openpyxl to access formatting
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Step 1: Unmerge cells (copy value to all merged cells)
            self._unmerge_cells(ws)
            
            # Step 2: Save unmerged version temporarily if needed
            if output_path:
                wb.save(output_path)
                logger.info(f"Preprocessed file saved to: {output_path}")
            
            # Step 3: Read with pandas (now that cells are unmerged)
            df = pd.read_excel(file_path, engine='openpyxl')
            
            # Step 4: Fix headers (detect and handle multi-level headers)
            df = self._fix_headers(df)
            
            # Step 5: Normalize column names
            df = self._normalize_columns(df)
            
            logger.info(f"Successfully preprocessed Excel file: {file_path}")
            return df
            
        except Exception as e:
            logger.warning(f"Preprocessing failed, using standard load: {e}")
            # Fallback to standard pandas read
            try:
                return pd.read_excel(file_path)
            except Exception as e2:
                logger.error(f"Standard load also failed: {e2}")
                return None
    
    def _unmerge_cells(self, ws):
        """
        Unmerge cells by copying value to all merged cells
        
        Args:
            ws: OpenPyXL worksheet object
        """
        merged_ranges = list(ws.merged_cells.ranges)
        
        if not merged_ranges:
            logger.debug("No merged cells found")
            return
        
        logger.info(f"Found {len(merged_ranges)} merged cell ranges")
        
        for merged_range in merged_ranges:
            # Get the top-left cell value
            top_left = merged_range.min_row, merged_range.min_col
            value = ws.cell(*top_left).value
            
            # Copy value to all cells in merged range
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row, col).value = value
            
            # Unmerge the cells
            ws.unmerge_cells(str(merged_range))
        
        logger.info(f"Unmerged {len(merged_ranges)} cell ranges")
    
    def _fix_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Fix multi-level headers by detecting and combining them
        
        Args:
            df: DataFrame with potential multi-level headers
            
        Returns:
            DataFrame with fixed headers
        """
        # Check if first row looks like a header (contains text, not all numbers)
        first_row = df.iloc[0] if len(df) > 0 else None
        
        if first_row is None:
            return df
        
        # Simple heuristic: if first row has mostly text values, it might be a header
        text_count = sum(1 for val in first_row if isinstance(val, str) and val.strip())
        
        # If first row has significant text, check if it should be combined with column names
        if text_count > len(df.columns) * 0.5:  # More than 50% text
            # Check if column names are generic (Unnamed, numbers, etc.)
            generic_cols = sum(1 for col in df.columns 
                             if str(col).startswith('Unnamed') or 
                                str(col).isdigit() or
                                'Column' in str(col))
            
            # If many generic columns, first row might be actual headers
            if generic_cols > len(df.columns) * 0.5:
                # Use first row as headers and drop it
                df.columns = [str(val).strip() if pd.notna(val) else f'Column_{i}' 
                             for i, val in enumerate(first_row)]
                df = df.iloc[1:].reset_index(drop=True)
                logger.info("Detected and fixed header row")
        
        return df
    
    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize column names: trim, remove newlines, fix spacing
        
        Args:
            df: DataFrame with columns to normalize
            
        Returns:
            DataFrame with normalized column names
        """
        new_columns = []
        for col in df.columns:
            col_str = str(col)
            # Remove newlines, extra spaces
            normalized = col_str.replace('\n', ' ').replace('\r', ' ')
            normalized = ' '.join(normalized.split())  # Collapse multiple spaces
            normalized = normalized.strip()
            
            # If empty after normalization, use default name
            if not normalized or normalized.lower() in ['nan', 'none', '']:
                normalized = f'Column_{len(new_columns) + 1}'
            
            new_columns.append(normalized)
        
        # Ensure unique column names
        seen = {}
        unique_columns = []
        for col in new_columns:
            if col in seen:
                seen[col] += 1
                unique_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                unique_columns.append(col)
        
        df.columns = unique_columns
        return df
    
    def preprocess_file(self, input_path: str, output_path: str) -> bool:
        """
        Preprocess file and save to output path
        
        Args:
            input_path: Input file path
            output_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            df = self.preprocess(input_path, output_path)
            if df is not None:
                # Save to output path
                df.to_excel(output_path, index=False, engine='openpyxl')
                logger.info(f"Preprocessed file saved to: {output_path}")
                return True
            return False
        except Exception as e:
            logger.error(f"Error preprocessing file: {e}", exc_info=True)
            return False

